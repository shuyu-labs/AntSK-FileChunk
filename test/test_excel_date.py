"""
测试Excel日期字段解析
"""
import sys
from pathlib import Path

# 添加父目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.antsk_filechunk.unified_document_parser import UnifiedDocumentParser
import openpyxl
import datetime

def create_test_excel():
    """创建一个测试用的Excel文件，包含日期字段"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试数据"
    
    # 添加表头
    ws['A1'] = "编号"
    ws['B1'] = "姓名"
    ws['C1'] = "出生日期"
    ws['D1'] = "入职时间"
    ws['E1'] = "工资"
    
    # 添加测试数据
    test_data = [
        (1, "张三", datetime.date(1990, 5, 15), datetime.datetime(2020, 3, 1, 9, 0, 0), 8500.50),
        (2, "李四", datetime.date(1985, 8, 22), datetime.datetime(2019, 6, 15, 9, 0, 0), 12000),
        (3, "王五", datetime.date(1992, 12, 3), datetime.datetime(2021, 1, 10, 9, 0, 0), 7800),
    ]
    
    for row_idx, (num, name, birth_date, hire_datetime, salary) in enumerate(test_data, start=2):
        ws.cell(row=row_idx, column=1, value=num)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=birth_date)
        ws.cell(row=row_idx, column=4, value=hire_datetime)
        ws.cell(row=row_idx, column=5, value=salary)
        
        # 设置日期格式
        ws.cell(row=row_idx, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=row_idx, column=4).number_format = 'yyyy-mm-dd hh:mm:ss'
    
    # 保存文件
    test_file = Path(__file__).parent / "test_excel_dates.xlsx"
    wb.save(test_file)
    print(f"✓ 测试Excel文件已创建: {test_file}")
    return test_file

def test_excel_date_parsing():
    """测试Excel日期解析"""
    print("=" * 60)
    print("测试Excel日期字段解析")
    print("=" * 60)
    
    # 创建测试文件
    test_file = create_test_excel()
    
    try:
        # 解析文件
        parser = UnifiedDocumentParser()
        print(f"\n正在解析文件: {test_file.name}")
        result = parser.parse_file(test_file)
        
        print(f"\n解析结果:")
        print(f"- 表格数量: {len(result.tables)}")
        print(f"- 图片数量: {len(result.images)}")
        
        # 检查表格内容
        if result.tables:
            print(f"\n表格内容:")
            for table in result.tables:
                print(f"\n工作表: {table.get('sheet_name', 'N/A')}")
                print(f"行数: {table['rows']}, 列数: {table['cols']}")
                
                print("\n表格数据:")
                for i, row in enumerate(table['data']):
                    print(f"  行 {i}: {row}")
                
                # 重点检查日期字段
                print("\n日期字段检查:")
                for i, row in enumerate(table['data']):
                    if i == 0:
                        continue  # 跳过表头
                    
                    birth_date = row[2] if len(row) > 2 else ""
                    hire_datetime = row[3] if len(row) > 3 else ""
                    
                    print(f"  第{i}行:")
                    print(f"    - 出生日期: {birth_date} (类型: {type(birth_date).__name__})")
                    print(f"    - 入职时间: {hire_datetime} (类型: {type(hire_datetime).__name__})")
                    
                    # 验证日期格式
                    if "1990-05-15" in birth_date or "1985-08-22" in birth_date or "1992-12-03" in birth_date:
                        print(f"    ✓ 出生日期格式正确")
                    else:
                        print(f"    ✗ 出生日期格式错误，预期为 YYYY-MM-DD 格式，实际为: {birth_date}")
                    
                    if "2020-03-01" in hire_datetime or "2019-06-15" in hire_datetime or "2021-01-10" in hire_datetime:
                        print(f"    ✓ 入职时间格式正确")
                    else:
                        print(f"    ✗ 入职时间格式错误，预期为 YYYY-MM-DD HH:MM:SS 格式，实际为: {hire_datetime}")
        
        # 输出Markdown内容
        print("\n" + "=" * 60)
        print("Markdown内容预览:")
        print("=" * 60)
        print(result.markdown_content[:500] if len(result.markdown_content) > 500 else result.markdown_content)
        
        print("\n" + "=" * 60)
        print("测试完成！")
        print("=" * 60)
        
        return True
        
    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # 清理测试文件（可选）
        # test_file.unlink(missing_ok=True)
        pass

if __name__ == "__main__":
    success = test_excel_date_parsing()
    sys.exit(0 if success else 1)

